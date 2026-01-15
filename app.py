# Nama fail: app.py
import streamlit as st
import pandas as pd
from fpdf import FPDF
import datetime
import io
import os # Tambahan baru untuk cek fail logo
import plotly.express as px
from supabase import create_client, Client
import openpyxl 

# --- TETAPAN HALAMAN (MESTI DI ATAS SEKALI) ---
st.set_page_config(layout="wide", page_title="Sistem Gaji Sawit", page_icon="🌴")

# ==============================================================================
# 1. SAMBUNGAN KE SUPABASE
# ==============================================================================
try:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    supabase: Client = create_client(url, key)
    
    # Ping database
    try:
        supabase.table('rekod_kos').select("id").limit(1).execute()
    except Exception:
        pass 

except KeyError:
    st.error("Ralat: Rahsia 'SUPABASE_URL' atau 'SUPABASE_KEY' tidak ditemui.")
    st.stop()
except Exception as e:
    print(f"Status Sambungan: {e}")

# ==============================================================================
# 2. FUNGSI LOG MASUK & KESELAMATAN
# ==============================================================================
def check_password():
    if "logged_in" in st.session_state and st.session_state["logged_in"] == True:
        return True
    try:
        correct_password = st.secrets["APP_PASSWORD"]
    except KeyError:
        st.error("Ralat: Rahsia 'APP_PASSWORD' tidak ditemui.")
        return False

    # Logo di halaman Login (Jika ada)
    if os.path.exists("logo.png"):
        col_img, col_txt = st.columns([1, 4])
        with col_img:
            st.image("logo.png", width=80)
        with col_txt:
            st.warning("🔒 Sila masukkan kata laluan.")
    else:
        st.warning("🔒 Sila masukkan kata laluan.")

    password = st.text_input("Kata Laluan:", type="password")

    if st.button("Log Masuk"):
        if password == correct_password:
            st.session_state["logged_in"] = True
            st.rerun()
        else:
            st.error("Kata laluan salah.")
    return False

if not check_password():
    st.stop()

# ==============================================================================
# 3. FUNGSI-FUNGSI LOGIK (KIRAAN, PDF, EXCEL)
# ==============================================================================

def kira_payroll(senarai_resit, total_kos):
    KADAR_LORI_PER_KG = 0.07
    jumlah_hasil_jualan = sum(resit['Hasil_RM'] for resit in senarai_resit)
    jumlah_berat_kg = sum(resit['Berat_kg'] for resit in senarai_resit)
    gaji_lori = jumlah_berat_kg * KADAR_LORI_PER_KG
    baki_bersih = jumlah_hasil_jualan - gaji_lori - total_kos
    gaji_penumbak = baki_bersih / 2
    bahagian_pemilik = baki_bersih / 2
    return {
        "jumlah_hasil_jualan": jumlah_hasil_jualan,
        "jumlah_berat_kg": jumlah_berat_kg,
        "jumlah_berat_mt": jumlah_berat_kg / 1000,
        "gaji_lori": gaji_lori,
        "total_kos_operasi": total_kos,
        "baki_bersih": baki_bersih,
        "gaji_penumbak": gaji_penumbak,
        "bahagian_pemilik": bahagian_pemilik,
        "kadar_lori_per_kg": KADAR_LORI_PER_KG
    }

def jana_pdf_binary(bulan_tahun, senarai_resit, data_kiraan):
    pdf = FPDF()
    pdf.add_page()
    
    # --- LOGO & HEADER ---
    # Jika logo wujud, letak di penjuru kiri atas
    if os.path.exists("logo.png"):
        # Image(nama_fail, x, y, width)
        pdf.image("logo.png", 10, 8, 25)
        # Anjakkan tajuk ke bawah sedikit supaya tidak bertindih logo
        pdf.set_y(35)
    else:
        pdf.set_y(20)

    pdf.set_font("Helvetica", 'B', 18)
    pdf.cell(0, 10, f"LADANG SAWIT SATIN LUNG MANIS", ln=True, align='C')
    pdf.set_font("Helvetica", 'B', 14)
    pdf.cell(0, 10, f"Laporan Kiraan Gaji - {bulan_tahun}", ln=True, align='C')
    pdf.ln(10)
    
    # Bahagian 1: Jualan
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 10, "Bahagian 1: Butiran Jualan (Resit)", ln=True)
    pdf.set_font("Helvetica", size=11)
    for i, resit in enumerate(senarai_resit):
        gred = resit.get('Gred', 'N/A')
        berat_kg = resit.get('Berat_kg', 0)
        harga_per_mt = resit.get('Harga_RM_per_MT', 0)
        hasil_resit = resit.get('Hasil_RM', 0)
        teks_resit = f"  Resit #{i+1} (Gred {gred}): {berat_kg:.2f} kg @ RM{harga_per_mt:.2f}/MT = RM{hasil_resit:.2f}"
        pdf.cell(0, 8, teks_resit, ln=True)
    pdf.ln(5)
    
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(0, 8, f"Jumlah Berat Keseluruhan: {data_kiraan.get('jumlah_berat_kg', 0):.2f} kg", ln=True)
    pdf.cell(0, 8, f"Jumlah Hasil Jualan Kasar: RM{data_kiraan.get('jumlah_hasil_jualan', 0):.2f}", ln=True)
    pdf.ln(10)

    # Bahagian 2: Kiraan Gaji
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 10, "Bahagian 2: Pengiraan Gaji dan Pembahagian", ln=True)
    
    # Gaji Lori
    pdf.set_font("Helvetica", 'BU', 11)
    pdf.cell(0, 8, "Gaji Pekerja 1 (Lori):", ln=True)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, f"  Kiraan: {data_kiraan.get('jumlah_berat_kg', 0):.2f} kg x RM{data_kiraan.get('kadar_lori_per_kg', 0.07):.2f}/kg", ln=True)
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(0, 8, f"  Jumlah Gaji Lori = RM{data_kiraan.get('gaji_lori', 0):.2f}", ln=True)
    pdf.ln(5)

    # Kos Operasi
    pdf.set_font("Helvetica", 'BU', 11)
    pdf.cell(0, 8, "Kos Operasi Bulanan (Baja, Racun, dll):", ln=True)
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(0, 8, f"  Jumlah Kos Operasi = RM{data_kiraan.get('total_kos_operasi', 0):.2f}", ln=True)
    pdf.ln(5)

    # Baki Bersih
    pdf.set_font("Helvetica", 'BU', 11)
    pdf.cell(0, 8, "Hasil Bersih (Untuk Dibahagi):", ln=True)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, f"  Kiraan: RM{data_kiraan.get('jumlah_hasil_jualan', 0):.2f} (Jualan) - RM{data_kiraan.get('gaji_lori', 0):.2f} (Lori) - RM{data_kiraan.get('total_kos_operasi', 0):.2f} (Kos Operasi)", ln=True)
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(0, 8, f"  Hasil Bersih = RM{data_kiraan.get('baki_bersih', 0):.2f}", ln=True)
    pdf.ln(5)

    # Pembahagian 50/50
    pdf.set_font("Helvetica", 'BU', 11)
    pdf.cell(0, 8, "Pembahagian Hasil Bersih (50/50):", ln=True)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(0, 8, f"  Kiraan: RM{data_kiraan.get('baki_bersih', 0):.2f} / 2", ln=True)
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(0, 8, f"  Gaji Pekerja 2 (Penumbak) = RM{data_kiraan.get('gaji_penumbak', 0):.2f}", ln=True)
    pdf.cell(0, 8, f"  Bahagian Pemilik Ladang = RM{data_kiraan.get('bahagian_pemilik', 0):.2f}", ln=True)
    pdf.ln(15)
    
    # Footer
    pdf.set_font("Helvetica", 'I', 9)
    pdf.cell(0, 5, "Laporan ini disediakan oleh:", ln=True, align='L')
    pdf.set_font("Helvetica", 'B', 9)
    pdf.cell(0, 5, "Mohamad Saifullah Satin", ln=True, align='L')
    pdf.set_font("Helvetica", 'I', 9)
    pdf.cell(0, 5, "Telefon: 019-840 6421", ln=True, align='L')
    pdf.cell(0, 5, "Email: msaifullahsatin@gmail.com", ln=True, align='L')
    tarikh_jana = datetime.date.today().strftime("%d-%m-%Y")
    pdf.set_y(-15)
    pdf.set_font("Helvetica", 'I', 8)
    nama_anda = st.secrets.get('NAMA_ANDA', 'Admin')
    pdf.cell(0, 10, f"Laporan dijana secara automatik pada {tarikh_jana} oleh {nama_anda}", ln=True, align='C')
    
    return bytes(pdf.output(dest='S'))

def jana_pdf_berkelompok(laporan_title, df_gaji_filtered, df_jualan_filtered, df_kos_filtered):
    pdf = FPDF()
    pdf.add_page()
    
    # --- LOGO & HEADER ---
    if os.path.exists("logo.png"):
        pdf.image("logo.png", 10, 8, 25)
        pdf.set_y(35)
    else:
        pdf.set_y(20)

    pdf.set_font("Helvetica", 'B', 18)
    pdf.cell(0, 10, f"LADANG SAWIT SATIN LUNG MANIS", ln=True, align='C')
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(0, 10, f"Ringkasan Laporan - {laporan_title}", ln=True, align='C')
    pdf.ln(10)

    # 2. Ringkasan Keseluruhan (KPI)
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 10, "Ringkasan Prestasi Keseluruhan", ln=True)
    
    total_jualan = df_gaji_filtered['JumlahJualan_RM'].sum()
    total_berat = df_gaji_filtered['JumlahBerat_kg'].sum()
    total_gaji_lori = df_gaji_filtered['GajiLori_RM'].sum()
    total_kos_ops = df_gaji_filtered.get('total_kos_operasi', 0.0).sum()
    total_gaji_penumbak = df_gaji_filtered['GajiPenumbak_RM'].sum()
    total_bahagian_pemilik = df_gaji_filtered['BahagianPemilik_RM'].sum()
    
    pdf.set_font("Helvetica", '', 11)
    pdf.cell(0, 8, f"Jumlah Jualan Kasar: RM {total_jualan:,.2f}", ln=True)
    pdf.cell(0, 8, f"Jumlah Berat Jualan: {total_berat:,.2f} kg", ln=True)
    pdf.cell(0, 8, f"Jumlah Kos Operasi: RM {total_kos_ops:,.2f}", ln=True)
    pdf.cell(0, 8, f"Jumlah Gaji Lori: RM {total_gaji_lori:,.2f}", ln=True)
    pdf.cell(0, 8, f"Jumlah Gaji Penumbak: RM {total_gaji_penumbak:,.2f}", ln=True)
    pdf.cell(0, 8, f"Jumlah Bahagian Pemilik: RM {total_bahagian_pemilik:,.2f}", ln=True)
    pdf.ln(10)

    # 3. Jadual Ringkasan Mengikut Bulan
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 10, "Pecahan Mengikut Bulan", ln=True)
    
    w_bulan = 40
    w_angka = 25 
    
    pdf.set_font("Helvetica", 'B', 8)
    pdf.cell(w_bulan, 8, "Bulan", 1, align='C')
    pdf.cell(w_angka, 8, "Jualan (RM)", 1, align='C')
    pdf.cell(w_angka, 8, "Kos Ops (RM)", 1, align='C')
    pdf.cell(w_angka, 8, "Gaji Lori (RM)", 1, align='C')
    pdf.cell(w_angka, 8, "Gaji Pnumbak (RM)", 1, align='C')
    pdf.cell(w_angka, 8, "Pemilik (RM)", 1, align='C')
    pdf.cell(w_angka, 8, "Berat (kg)", 1, ln=True, align='C')

    pdf.set_font("Helvetica", '', 8)
    try:
        peta_bulan = {
            "Januari": 1, "Februari": 2, "Mac": 3, "April": 4, "Mei": 5, "Jun": 6,
            "Julai": 7, "Ogos": 8, "September": 9, "Oktober": 10, "November": 11, "Disember": 12
        }
        df_gaji_filtered['BulanString'] = df_gaji_filtered['BulanTahun'].str.split(' ', expand=True)[0]
        df_gaji_filtered['Tahun'] = df_gaji_filtered['BulanTahun'].str.split(' ', expand=True)[1].astype(int)
        df_gaji_filtered['BulanNombor'] = df_gaji_filtered['BulanString'].map(peta_bulan)
        df_gaji_sorted = df_gaji_filtered.sort_values(by=['Tahun', 'BulanNombor'])
    except Exception:
        df_gaji_sorted = df_gaji_filtered
        
    for index, data in df_gaji_sorted.iterrows():
        pdf.cell(w_bulan, 8, data['BulanTahun'], 1)
        pdf.cell(w_angka, 8, f"{data['JumlahJualan_RM']:,.2f}", 1, align='R')
        pdf.cell(w_angka, 8, f"{data.get('total_kos_operasi', 0.0):,.2f}", 1, align='R')
        pdf.cell(w_angka, 8, f"{data['GajiLori_RM']:,.2f}", 1, align='R')
        pdf.cell(w_angka, 8, f"{data['GajiPenumbak_RM']:,.2f}", 1, align='R')
        pdf.cell(w_angka, 8, f"{data['BahagianPemilik_RM']:,.2f}", 1, align='R')
        pdf.cell(w_angka, 8, f"{data['JumlahBerat_kg']:,.2f}", 1, ln=True, align='R')
    
    pdf.ln(10)
    
    # 4. Ringkasan Pecahan Jualan (Gred) & Kos
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 10, "Pecahan Keseluruhan (Gred & Kos)", ln=True)
    pdf.set_font("Helvetica", '', 11)

    # Pecahan Gred
    gred_a_berat = df_jualan_filtered[df_jualan_filtered['Gred'] == 'A']['Berat_kg'].sum()
    gred_a_hasil = df_jualan_filtered[df_jualan_filtered['Gred'] == 'A']['Hasil_RM'].sum()
    gred_b_berat = df_jualan_filtered[df_jualan_filtered['Gred'] == 'B']['Berat_kg'].sum()
    gred_b_hasil = df_jualan_filtered[df_jualan_filtered['Gred'] == 'B']['Hasil_RM'].sum()
    gred_c_berat = df_jualan_filtered[df_jualan_filtered['Gred'] == 'C']['Berat_kg'].sum()
    gred_c_hasil = df_jualan_filtered[df_jualan_filtered['Gred'] == 'C']['Hasil_RM'].sum()
    
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(95, 8, "Pecahan Jualan (Gred)", 1, ln=True, align='C')
    pdf.set_font("Helvetica", '', 10)
    pdf.cell(0, 8, f"Gred A: {gred_a_berat:,.2f} kg  |  RM {gred_a_hasil:,.2f}", ln=True)
    pdf.cell(0, 8, f"Gred B: {gred_b_berat:,.2f} kg  |  RM {gred_b_hasil:,.2f}", ln=True)
    pdf.cell(0, 8, f"Gred C: {gred_c_berat:,.2f} kg  |  RM {gred_c_hasil:,.2f}", ln=True)
    pdf.ln(5)

    # Pecahan Kos
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(95, 8, "Pecahan Kos Operasi", 1, ln=True, align='C')
    pdf.set_font("Helvetica", '', 10)
    if not df_kos_filtered.empty:
        kos_by_type = df_kos_filtered.groupby('JenisKos')['Jumlah_RM'].sum()
        for jenis, jumlah in kos_by_type.items():
            pdf.cell(0, 8, f"{jenis}: RM {jumlah:,.2f}", ln=True)
    else:
        pdf.cell(0, 8, "Tiada kos operasi direkodkan.", ln=True)
    
    # 5. Footer
    tarikh_jana = datetime.date.today().strftime("%d-%m-%Y")
    pdf.set_y(-15)
    pdf.set_font("Helvetica", 'I', 8)
    nama_anda = st.secrets.get('NAMA_ANDA', 'Admin')
    pdf.cell(0, 10, f"Laporan dijana secara automatik pada {tarikh_jana} oleh {nama_anda}", ln=True, align='C')
    
    return bytes(pdf.output(dest='S'))

def to_excel(df_gaji, df_jualan, df_kos):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_gaji.to_excel(writer, sheet_name='Ringkasan_Gaji', index=False)
        df_jualan.to_excel(writer, sheet_name='Butiran_Jualan', index=False)
        df_kos.to_excel(writer, sheet_name='Butiran_Kos', index=False)
    processed_data = output.getvalue()
    return processed_data

def proses_dataframe_bulanan(df_gaji_raw):
    if df_gaji_raw.empty:
        return pd.DataFrame(columns=['BulanTahun', 'Tahun', 'BulanNombor', 'BulanString', 'JumlahJualan_RM', 'total_kos_operasi', 'Keuntungan_RM'])

    df = df_gaji_raw.copy()
    if 'total_kos_operasi' not in df.columns:
        df['total_kos_operasi'] = 0.0
    df['total_kos_operasi'] = df['total_kos_operasi'].fillna(0)

    df['Keuntungan_RM'] = df['JumlahJualan_RM'] - df['GajiLori_RM'] - df['total_kos_operasi']

    try:
        peta_bulan = {
            "Januari": 1, "Februari": 2, "Mac": 3, "April": 4, "Mei": 5, "Jun": 6,
            "Julai": 7, "Ogos": 8, "September": 9, "Oktober": 10, "November": 11, "Disember": 12
        }
        df_split = df['BulanTahun'].str.split(' ', expand=True)
        df['BulanString'] = df_split[0]
        df['Tahun'] = df_split[1].astype(int)
        df['BulanNombor'] = df['BulanString'].map(peta_bulan)
    except Exception:
        df['Tahun'] = 2000
        df['BulanNombor'] = 1
        df['BulanString'] = 'N/A'

    return df

# ==============================================================================
# 4. MUATKAN DATA
# ==============================================================================
@st.cache_data(ttl=600)
def muat_data():
    try:
        response_gaji = supabase.table('rekod_gaji').select("*").order('id', desc=False).execute()
        df_gaji = pd.DataFrame(response_gaji.data)
        
        response_jualan = supabase.table('rekod_jualan').select("*").order('id', desc=False).execute()
        df_jualan = pd.DataFrame(response_jualan.data)
        
        response_kos = supabase.table('rekod_kos').select("*").order('id', desc=False).execute()
        df_kos = pd.DataFrame(response_kos.data)
        
        expected_gaji_cols = ['BulanTahun', 'JumlahJualan_RM', 'JumlahBerat_kg', 'GajiLori_RM', 
                              'GajiPenumbak_RM', 'BahagianPemilik_RM', 'total_kos_operasi', 'id', 'created_at']
        expected_jualan_cols = ['BulanTahun', 'IDResit', 'Gred', 'Berat_kg', 'Harga_RM_per_MT', 'Hasil_RM', 'id', 'created_at']
        expected_kos_cols = ['BulanTahun', 'JenisKos', 'Jumlah_RM', 'id', 'created_at']

        if df_gaji.empty: df_gaji = pd.DataFrame(columns=expected_gaji_cols)
        if df_jualan.empty: df_jualan = pd.DataFrame(columns=expected_jualan_cols)
        if df_kos.empty: df_kos = pd.DataFrame(columns=expected_kos_cols)

        return df_gaji, df_jualan, df_kos
    except Exception as e:
        st.error(f"Ralat database: {e}")
        return (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())

df_gaji_raw, df_jualan_raw, df_kos_raw = muat_data()
df_gaji_processed = proses_dataframe_bulanan(df_gaji_raw)

# ==============================================================================
# 5. PAPARAN UTAMA
# ==============================================================================
# --- LOGO DI SIDEBAR ---
if os.path.exists("logo.png"):
    st.sidebar.image("logo.png", use_column_width=True)

st.title("Sistem Pengurusan Ladang Sawit 🧑‍🌾")

st.sidebar.title("Navigasi")
page = st.sidebar.radio("Pilih Halaman:", ["📊 Dashboard Statistik", 
                                          "📝 Kemasukan Data Baru", 
                                          "🖨️ Urus & Cetak Semula", 
                                          "📈 Laporan Berkelompok"])

if st.sidebar.button("Segarkan Semula Data (Refresh)"):
    st.cache_data.clear()
    st.rerun()

st.sidebar.error("Klik untuk keluar dari sistem.")
if st.sidebar.button("Log Keluar"):
    st.session_state["logged_in"] = False
    st.rerun()

# --- HALAMAN 1: DASHBOARD ---
if page == "📊 Dashboard Statistik":
    st.header("📊 Dashboard Statistik")
    
    tab_tren, tab_perbandingan = st.tabs(["📈 Tren Keseluruhan", "⚖️ Perbandingan Tahun-ke-Tahun"])

    with tab_tren:
        if df_gaji_processed.empty:
            st.warning("Tiada data. Sila masukkan data baru.")
        else:
            total_sales = df_gaji_processed['JumlahJualan_RM'].sum()
            total_weight_kg = df_gaji_processed['JumlahBerat_kg'].sum()
            avg_monthly_owner = df_gaji_processed['BahagianPemilik_RM'].mean()

            c1, c2, c3 = st.columns(3)
            c1.metric("Jualan Keseluruhan", f"RM{total_sales:,.2f}")
            c2.metric("Berat Keseluruhan", f"{total_weight_kg:,.0f} kg")
            c3.metric("Purata Pemilik (Bulanan)", f"RM{avg_monthly_owner:,.2f}")
            st.markdown("---")
            
            st.subheader("Tren Jualan, Kos & Untung")
            df_srt = df_gaji_processed.sort_values(by=['Tahun', 'BulanNombor'])
            fig_tren = px.line(df_srt, x='BulanTahun', y=['JumlahJualan_RM', 'total_kos_operasi', 'Keuntungan_RM', 'BahagianPemilik_RM'], markers=True)
            st.plotly_chart(fig_tren, use_container_width=True)
            
            st.subheader("Analisis Pecahan")
            cg1, cg2 = st.columns(2)
            with cg1:
                if not df_jualan_raw.empty:
                    fig_p = px.pie(df_jualan_raw, names='Gred', values='Hasil_RM', title="Pecahan Jualan (Gred)")
                    st.plotly_chart(fig_p, use_container_width=True)
            with cg2:
                if not df_kos_raw.empty and df_kos_raw['Jumlah_RM'].sum() > 0:
                    fig_k = px.pie(df_kos_raw, names='JenisKos', values='Jumlah_RM', title="Pecahan Kos")
                    st.plotly_chart(fig_k, use_container_width=True)
                else:
                    st.info("Tiada rekod kos.")

    with tab_perbandingan:
        st.subheader("Perbandingan Tahun-ke-Tahun")
        yrs = sorted(df_gaji_processed['Tahun'].unique(), reverse=True)
        if len(yrs) < 2:
            st.info("Perlukan min 2 tahun data.")
        else:
            cy1, cy2 = st.columns(2)
            y1 = cy1.selectbox("Tahun 1:", yrs, index=1)
            y2 = cy2.selectbox("Tahun 2:", yrs, index=0)

            if y1 == y2:
                st.error("Pilih tahun berbeza.")
            else:
                bulan_inv = {1: "Jan", 2: "Feb", 3: "Mac", 4: "Apr", 5: "Mei", 6: "Jun", 7: "Jul", 8: "Ogos", 9: "Sep", 10: "Okt", 11: "Nov", 12: "Dis"}
                
                d1 = df_gaji_processed[df_gaji_processed['Tahun']==y1][['BulanNombor','JumlahJualan_RM','Keuntungan_RM']].add_suffix(f"_{y1}").rename(columns={f'BulanNombor_{y1}':'BN'})
                d2 = df_gaji_processed[df_gaji_processed['Tahun']==y2][['BulanNombor','JumlahJualan_RM','Keuntungan_RM']].add_suffix(f"_{y2}").rename(columns={f'BulanNombor_{y2}':'BN'})
                
                dm = pd.merge(d1, d2, on='BN', how='outer').fillna(0)
                dm['Bulan'] = dm['BN'].map(bulan_inv)
                dm.sort_values('BN', inplace=True)

                fig_j = px.bar(dm, x='Bulan', y=[f'JumlahJualan_RM_{y1}', f'JumlahJualan_RM_{y2}'], barmode='group', title="Perbandingan Jualan")
                st.plotly_chart(fig_j, use_container_width=True)

# --- HALAMAN 2: KEMASUKAN DATA ---
elif page == "📝 Kemasukan Data Baru":
    st.header("📝 Kemasukan Data Baru")
    tj, tk = st.tabs(["1. Jualan (Gaji)", "2. Kos Operasi"])
    bln_list = ["Januari", "Februari", "Mac", "April", "Mei", "Jun", "Julai", "Ogos", "September", "Oktober", "November", "Disember"]
    thn_curr = datetime.date.today().year
    thn_list = list(range(thn_curr - 5, thn_curr + 2))[::-1]

    with tj:
        st.subheader("Borang Gaji")
        with st.form("f_gaji"):
            c1, c2 = st.columns(2)
            bg = c1.selectbox("Bulan:", bln_list, index=datetime.date.today().month-1, key="bg")
            tg = c2.selectbox("Tahun:", thn_list, key="tg")
            bt_gaji = f"{bg} {tg}"
            st.info(f"Untuk: **{bt_gaji}**")
            
            df_in = pd.DataFrame([{"Gred": "A", "Berat_kg": 0.0, "Harga_RM_per_MT": 0.0}, {"Gred": "B", "Berat_kg": 0.0, "Harga_RM_per_MT": 0.0}])
            ed_j = st.data_editor(df_in, num_rows="dynamic", column_config={"Gred": st.column_config.SelectboxColumn("Gred", options=["A","B","C"])})
            sub_g = st.form_submit_button("Simpan Gaji")

    with tk:
        st.subheader("Borang Kos")
        with st.form("f_kos"):
            c1, c2 = st.columns(2)
            bk = c1.selectbox("Bulan:", bln_list, index=datetime.date.today().month-1, key="bk")
            tkos = c2.selectbox("Tahun:", thn_list, key="tk")
            bt_kos = f"{bk} {tkos}"
            st.info(f"Untuk: **{bt_kos}**")
            
            df_k_in = pd.DataFrame([{"JenisKos": "Baja", "Jumlah_RM": 0.0}])
            ed_k = st.data_editor(df_k_in, num_rows="dynamic")
            sub_k = st.form_submit_button("Simpan Kos")

    if sub_k:
        if ed_k['Jumlah_RM'].sum() == 0: st.error("Masukan kos > 0")
        else:
            lk = ed_k[ed_k['Jumlah_RM']>0].to_dict('records')
            for k in lk: k['BulanTahun'] = bt_kos
            try:
                if not df_kos_raw.empty: supabase.table('rekod_kos').delete().eq('BulanTahun', bt_kos).execute()
                supabase.table('rekod_kos').insert(lk).execute()
                st.cache_data.clear()
                st.success("Kos disimpan!")
            except Exception as e: st.error(str(e))

    if sub_g:
        if ed_j['Berat_kg'].sum() == 0: st.error("Tiada resit dimasukkan.")
        elif not df_gaji_raw.empty and bt_gaji in df_gaji_raw['BulanTahun'].values: st.error("Data wujud.")
        else:
            kos_semasa = df_kos_raw[df_kos_raw['BulanTahun']==bt_gaji]['Jumlah_RM'].sum() if not df_kos_raw.empty else 0.0
            l_res = ed_j[ed_j['Berat_kg']>0].to_dict('records')
            for i, r in enumerate(l_res):
                r['Hasil_RM'] = (r['Berat_kg']/1000)*r['Harga_RM_per_MT']
                r['BulanTahun'] = bt_gaji
                r['IDResit'] = i+1
            
            dat = kira_payroll(l_res, kos_semasa)
            pdf = jana_pdf_binary(bt_gaji, l_res, dat)
            
            dg = {'BulanTahun': bt_gaji, 'JumlahJualan_RM': dat['jumlah_hasil_jualan'], 'JumlahBerat_kg': dat['jumlah_berat_kg'], 'GajiLori_RM': dat['gaji_lori'], 'GajiPenumbak_RM': dat['gaji_penumbak'], 'BahagianPemilik_RM': dat['bahagian_pemilik'], 'total_kos_operasi': dat['total_kos_operasi']}
            dj = [{'BulanTahun': r['BulanTahun'], 'IDResit': r['IDResit'], 'Gred': r['Gred'], 'Berat_kg': r['Berat_kg'], 'Harga_RM_per_MT': r['Harga_RM_per_MT'], 'Hasil_RM': r['Hasil_RM']} for r in l_res]

            try:
                supabase.table('rekod_gaji').insert(dg).execute()
                supabase.table('rekod_jualan').insert(dj).execute()
                st.cache_data.clear()
                st.success("Gaji disimpan!")
                st.download_button("Download PDF", pdf, f"Laporan_{bt_gaji}.pdf", "application/pdf")
            except Exception as e: st.error(str(e))

# --- HALAMAN 3: URUS & CETAK ---
elif page == "🖨️ Urus & Cetak Semula":
    st.header("🖨️ Urus & Cetak Semula")
    if df_gaji_raw.empty: st.info("Tiada data.")
    else:
        sb = df_gaji_raw['BulanTahun'].unique()
        
        st.subheader("1. Cetak Semula")
        c1, c2 = st.columns([3,1])
        bc = c1.selectbox("Pilih Bulan:", sb)
        if bc:
            dg = df_gaji_raw[df_gaji_raw['BulanTahun']==bc].to_dict('records')[0]
            lr = df_jualan_raw[df_jualan_raw['BulanTahun']==bc].to_dict('records')
            dt = {'jumlah_hasil_jualan': dg['JumlahJualan_RM'], 'jumlah_berat_kg': dg['JumlahBerat_kg'], 'gaji_lori': dg['GajiLori_RM'], 'total_kos_operasi': dg.get('total_kos_operasi',0.0), 'kadar_lori_per_kg': 0.07, 'baki_bersih': dg['GajiPenumbak_RM']+dg['BahagianPemilik_RM'], 'gaji_penumbak': dg['GajiPenumbak_RM'], 'bahagian_pemilik': dg['BahagianPemilik_RM']}
            pdf = jana_pdf_binary(bc, lr, dt)
            c2.write(" ")
            c2.write(" ")
            c2.download_button("Download PDF", pdf, f"Laporan_{bc}.pdf", "application/pdf")
        
        st.divider()
        st.subheader("2. Edit Data")
        if 'be' not in st.session_state: st.session_state.be = None
        c1, c2 = st.columns([3,1])
        be = c1.selectbox("Bulan Edit:", sb, key="sbe")
        if c2.button("Load"): st.session_state.be = be; st.rerun()

        if st.session_state.be:
            ba = st.session_state.be
            st.warning(f"Edit: **{ba}**")
            dj = df_jualan_raw[df_jualan_raw['BulanTahun']==ba][['Gred','Berat_kg','Harga_RM_per_MT']]
            dk = df_kos_raw[df_kos_raw['BulanTahun']==ba][['JenisKos','Jumlah_RM']]
            
            with st.form("fe"):
                st.write("Jualan:")
                ej = st.data_editor(dj, num_rows="dynamic")
                st.write("Kos:")
                ek = st.data_editor(dk, num_rows="dynamic")
                if st.form_submit_button("Simpan"):
                    try:
                        supabase.table('rekod_gaji').delete().eq('BulanTahun', ba).execute()
                        supabase.table('rekod_jualan').delete().eq('BulanTahun', ba).execute()
                        supabase.table('rekod_kos').delete().eq('BulanTahun', ba).execute()
                        
                        kb = 0.0
                        if not ek.empty and ek['Jumlah_RM'].sum()>0:
                            lk = ek[ek['Jumlah_RM']>0].to_dict('records')
                            for k in lk: k['BulanTahun'] = ba
                            supabase.table('rekod_kos').insert(lk).execute()
                            kb = sum(k['Jumlah_RM'] for k in lk)
                        
                        if not ej.empty and ej['Berat_kg'].sum()>0:
                            lr = ej[ej['Berat_kg']>0].to_dict('records')
                            for i,r in enumerate(lr):
                                r['Hasil_RM'] = (r['Berat_kg']/1000)*r['Harga_RM_per_MT']
                                r['BulanTahun'] = ba
                                r['IDResit'] = i+1
                            da = kira_payroll(lr, kb)
                            dg = {'BulanTahun': ba, 'JumlahJualan_RM': da['jumlah_hasil_jualan'], 'JumlahBerat_kg': da['jumlah_berat_kg'], 'GajiLori_RM': da['gaji_lori'], 'GajiPenumbak_RM': da['gaji_penumbak'], 'BahagianPemilik_RM': da['bahagian_pemilik'], 'total_kos_operasi': da['total_kos_operasi']}
                            dj2 = [{'BulanTahun': r['BulanTahun'], 'IDResit': r['IDResit'], 'Gred': r['Gred'], 'Berat_kg': r['Berat_kg'], 'Harga_RM_per_MT': r['Harga_RM_per_MT'], 'Hasil_RM': r['Hasil_RM']} for r in lr]
                            supabase.table('rekod_gaji').insert(dg).execute()
                            supabase.table('rekod_jualan').insert(dj2).execute()

                        st.cache_data.clear()
                        st.session_state.be = None
                        st.success("Updated!")
                        st.rerun()
                    except Exception as e: st.error(str(e))

        st.divider()
        st.subheader("3. Padam Data")
        with st.form("fd"):
            bd = st.selectbox("Padam Bulan:", sb)
            if st.form_submit_button("Padam Kekal"):
                supabase.table('rekod_gaji').delete().eq('BulanTahun', bd).execute()
                supabase.table('rekod_jualan').delete().eq('BulanTahun', bd).execute()
                supabase.table('rekod_kos').delete().eq('BulanTahun', bd).execute()
                st.cache_data.clear()
                st.success("Deleted.")
                st.rerun()

        st.divider()
        st.subheader("4. Backup")
        exc = to_excel(df_gaji_raw, df_jualan_raw, df_kos_raw)
        st.download_button("Download Excel", exc, f"backup_{datetime.date.today()}.xlsx")

# --- HALAMAN 4: LAPORAN ---
elif page == "📈 Laporan Berkelompok":
    st.header("📈 Laporan")
    if df_gaji_processed.empty: st.info("Tiada data.")
    else:
        yrs = sorted(df_gaji_processed['Tahun'].unique(), reverse=True)
        with st.form("fr"):
            y = st.selectbox("Tahun:", yrs)
            ty = st.radio("Jenis:", ["Separuh 1 (Jan-Jun)", "Separuh 2 (Jul-Dis)", "Penuh"])
            if st.form_submit_button("Jana PDF"):
                m1 = ["Januari", "Februari", "Mac", "April", "Mei", "Jun"]
                m2 = ["Julai", "Ogos", "September", "Oktober", "November", "Disember"]
                if "1" in ty: ml, tt = [f"{b} {y}" for b in m1], f"Separuh 1 {y}"
                elif "2" in ty: ml, tt = [f"{b} {y}" for b in m2], f"Separuh 2 {y}"
                else: ml, tt = [f"{b} {y}" for b in m1+m2], f"Penuh {y}"
                
                d1 = df_gaji_raw[df_gaji_raw['BulanTahun'].isin(ml)]
                d2 = df_jualan_raw[df_jualan_raw['BulanTahun'].isin(ml)]
                d3 = df_kos_raw[df_kos_raw['BulanTahun'].isin(ml)]
                
                if d1.empty: st.error("Tiada data.")
                else:
                    pdf = jana_pdf_berkelompok(tt, d1, d2, d3)
                    st.download_button("Download PDF", pdf, f"Laporan_{tt}.pdf")
